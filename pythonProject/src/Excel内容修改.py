import openpyxl

wd = openpyxl.load_workbook('test.xlsx')  # 加载文件
sh = wd['Sheet']  # 获取sheet
sh.cell(2, 1, '修改后的用户名')  # 修改指定单元格的内容
print(sh.max_row)  # 获取行数
print(sh.max_column)  # 列数
print(sh.cell(1, 2).value)  # 获取1行2列的值
for row in sh.rows:
    for v in row:
        print(v.value)
wd.save('test.xlsx')  # 保存
wd.close()  # 关闭文件
