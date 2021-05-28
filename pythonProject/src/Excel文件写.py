import openpyxl

# 创建文件对象
wd = openpyxl.Workbook()
# 打开文件并激活当前的shell页
sh = wd.active
# sh.cell(第几行, 第几列, 写入的数据)
sh.cell(1, 1, '用户名')
sh.cell(1, 2, '密码')
accounts = [['user1', 'passwd1'], ['user2', 'passwd2'], ['user3', 'passwd3']]
row = 2
col = 1
for account in accounts:  # 循环外层，取一行数据
    for val in account:  # 控制写如表格中
        sh.cell(row, col, val)  # 写入第row行，第col列
        col = col + 1  # 下次写第二列
    col = 1  # 回到第一列
    row += 1  # 下一行
# 保存文件
wd.save('test88.xlsx')
wd.close()  # 关闭对象
